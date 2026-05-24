"""Inspeciona os 2 arquivos Excel enviados pela Ford."""
import sys
# Force UTF-8 stdout on Windows
sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")
__doc__ = """Inspeciona os 2 arquivos Excel enviados pela Ford.
- Lista abas, colunas, contagem de linhas, tipos e amostra.
- Salva resumo JSON pra ser usado em decisões de pipeline.
"""
import json
import sys
from pathlib import Path
import pandas as pd
import openpyxl

import os
BASE = os.environ.get("TEMP", "C:/Users/pedro.martins/AppData/Local/Temp") + "/ford-rar"
FILES = {
    "d1": f"{BASE}/FIAP-Ford - Data sheet_Desafio_01_v02.xlsx",
    "d2": f"{BASE}/vin_share_Desafio_02.xlsx",
}

def inspect(label: str, path: str) -> dict:
    print(f"\n===== {label}: {Path(path).name} =====")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        print(f"Abas: {sheets}")
        wb.close()
    except Exception as e:
        print("ERROR loading workbook:", e)
        return {}

    report = {"file": Path(path).name, "sheets": {}}
    for sheet in sheets:
        try:
            # Lê só 50 linhas pra amostra + pega total via openpyxl
            df = pd.read_excel(path, sheet_name=sheet, nrows=5)
            wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb2[sheet]
            total_rows = ws.max_row - 1  # menos header
            total_cols = ws.max_column
            wb2.close()

            print(f"\n  📋 Aba '{sheet}' — {total_rows} linhas × {total_cols} colunas")
            print(f"     Colunas: {list(df.columns)[:30]}")
            print(f"     Dtypes:")
            for col, dt in df.dtypes.items():
                print(f"        {col}: {dt}")
            print(f"     Amostra (primeiras 3 linhas):")
            print(df.head(3).to_string(max_colwidth=30))

            report["sheets"][sheet] = {
                "total_rows": int(total_rows),
                "total_cols": int(total_cols),
                "columns": [str(c) for c in df.columns],
                "dtypes": {str(c): str(d) for c, d in df.dtypes.items()},
                "sample": df.head(3).fillna("").astype(str).to_dict(orient='records'),
            }
        except Exception as e:
            print(f"  ❌ Erro na aba '{sheet}':", e)
            report["sheets"][sheet] = {"error": str(e)}
    return report

if __name__ == "__main__":
    out = {}
    for label, path in FILES.items():
        if Path(path).exists():
            out[label] = inspect(label, path)
    Path("C:/Users/pedro.martins/AppData/Local/Temp/ford-xlsx-inspection.json").write_text(
        json.dumps(out, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print("\n\nResumo salvo em C:/Users/pedro.martins/AppData/Local/Temp/ford-xlsx-inspection.json")
